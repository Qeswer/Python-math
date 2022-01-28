import openpyxl
import re
import matplotlib.pyplot as plt

exel_file = openpyxl.load_workbook("value.xlsx")
book1 = exel_file.active


def y_list(x, col):  # функция вытаскивания значений в столбик
    y_list_ = []
    for i in range(1, x.max_row):
        y_list_.append(x.cell(row=i, column=col).value)
    # print(y_list_)
    return y_list_[1:len(y_list_)]


time = y_list(book1, 1)


# print(time)

def reg(x):  # функция регулярного выражения
    list1 = []
    list2 = []
    for i in range(len(x)):
        list1.append(re.sub(r'2021-11-29T12:', '', x[i]))
        list2.append(re.findall(r'\d{2}:\d{2}', list1[i]))
        list1[i] = list2[i][0]
    return list1


time = reg(time)


# print(reg(time))


def sensors():  # список всех показаний датчиков
    sensors = []
    for i in range(2, book1.max_column):
        sensors.append(y_list(book1, i))
    return sensors


sensors_ = sensors()


# a = sensors_[0][1]-sensors_[0][0]
# print(len(sensors_))
# print(a)
# VBRYP1 = y_list(book1,2)
# VBRYP2 = y_list(book1,3)
# VBRYP3 = y_list(book1,4)
# VBRYP4 = y_list(book1,5)
# VBRYP5 = y_list(book1,6)
# VBRYP6 = y_list(book1,7)
# VBRYP7 = y_list(book1,8)
# VBRYP8 = y_list(book1,9)
# VBRYP9 = y_list(book1,10)
# VBRYP10 = y_list(book1,11)
# print((VBRYP1))

def delta(sens):
    list1 = [[0] * len(sens[0])] * len(sens)
    list2 = []
    list3 = []
    # print(len(list2))
    # print(len(sens))
    # print(sens[0][1])
    # print(len(sens[0]))
    for i in range(len(list1)):
        for j in range(len(list1[i])):
            a = sens[i][j];
            b = sens[i][0];
            c = a - b;
            list2.append(c)
            # a = sens[i][j]; b = sens[i][0]
            # list1[i][j] = a - b
        list3.append(list2)
        list2 = list3
    ___print___(list3[1])
    return list1


def ___print___(x):
    for i in range(len(x)):
        print(x[i])

def _print_(x):
    for i in x:
        print()
        for j in i:
            print(j, end=" ")


sensor = delta(sensors_)
# print(sensor[1], sensor[2])
# _print_(sensor)
# print(sensor[0])
# print(sensor[1])
# print(len(sensor))

# def plot(time,sensors):
#     print(sensors)
#     for i in range(len(sensors)):
#         plt.plot(time, sensors[i])
#     plt.show()
#
#
# graf = plot(time,sensor)
# plt.show()


# print(delta(sensors_)[1])
# print(len(delta(sensors_)))
# DVBRYP1 = delta(VBRYP1)
# DVBRYP2 = delta(VBRYP2)
# DVBRYP3 = delta(VBRYP3)
# DVBRYP4 = delta(VBRYP4)
# DVBRYP5 = delta(VBRYP5)
# DVBRYP6 = delta(VBRYP6)
# DVBRYP7 = delta(VBRYP7)
# DVBRYP8 = delta(VBRYP8)
# DVBRYP9 = delta(VBRYP9)
# DVBRYP10 = delta(VBRYP10)

# print(DVBRYP1)

# plt.plot(time,DVBRYP1)
# plt.plot(time,DVBRYP2)
# plt.plot(time,DVBRYP3)
# plt.plot(time,DVBRYP4)
# plt.plot(time,DVBRYP5)
# plt.plot(time,DVBRYP6)
# plt.plot(time,DVBRYP7)
# plt.plot(time,DVBRYP8)
# plt.plot(time,DVBRYP9)
# plt.plot(time,DVBRYP10)

# plt.xlabel('time')
# plt.ylabel('Нм')
# plt.title('Зависимость длины волны от времени')
# # plt.legend(['ВБР УП 1'])
#
# plt.show()


# book1 = exel_file["разница"]
# x_list = []
# for i in range(1, book1.max_column+1):
#     x_list.append(book1.cell(row=3, column=i).value) # все элементы строки
# print(x_list)

# book1 = exel_file["разница"]
# y_list = []
# for i in range(1, book1.max_row+1):
#     y_list.append(book1.cell(row=i, column=2).value) # вытащим все элементы 2-го столбца
# del y_list[1] #удалим 2 элемент листа
# print(y_list)

# book1 = exel_file['максимумы']
# print(f'rows = {book1.max_row} \ncolumns = {book1.max_column}') #общее количество строк и столбцов

# book1 = exel_file['разница']
# print(f'лист3 {book1.cell(row=1, column=2).value}') #значение на листе разница

# cell_obj = currently_active_sheet.cell(row=2, column=2) //row>1
# print(f'time={cell_obj.value}') # вывод информации из определенной ячейки
# print(f'название листов{exel_file.sheetnames}')

# черновик
# def exel():
#     sensors = openpyxl.open("value.xlsx", read_only=True).active()
#     sell = sensors["A1":"A15"]
#
#     col: int
#     for col in range(1, sell):
#         print(col)
#
#
# exel()
# book = openpyxl.Workbook()
#
# sheet = book.active

# sheet["A1"] = 100
# sheet["B2"] = "world"
#
# sheet[1][1].value = 200
#
# book.save("book.xlsx")
# book.close()

# # СЧИТЫВАНИЕ ДАННЫХ ЭКСЕЛЬ
# book = openpyxl.open('value.xlsx', read_only=True)
#
# sheet = book.active
# cells = sheet["B2":"C12"]
# for sens1, sens2 in cells:
#     print(sens1.value, sens2.value)

# print(sheet[2][1].value)
# for row in range(1,sheet.max_row):
#     time = sheet[row][0].value
#     print(row, time)

# for row in sheet.iter_rows(min_row=2, max_row=15, min_col=1, max_col=3):
#     for cell in row:
#         print(cell.value, end=" ")
#     print()
