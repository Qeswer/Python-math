import openpyxl
import re
import matplotlib.pyplot as plt
from matplotlib.pyplot import MultipleLocator
import matplotlib.ticker as ticker
import numpy as np

exel_file = openpyxl.load_workbook("value.xlsx")
book2 = exel_file['разница']


def y_list(x, col):  # функция вытаскивания значений в столбик
    y_list_ = []
    for i in range(1, x.max_row):
        y_list_.append(x.cell(row=i, column=col).value)
    # print(y_list_)
    return y_list_[1:len(y_list_)]


def name_sens(x, row_num):
    sensors_name = []
    for i in range(2, x.max_column):
        sensors_name.append(x.cell(row=row_num, column=i).value)
    return sensors_name


def reg(x):  # функция регулярного выражения
    list1 = []
    list2 = []
    for i in range(len(x)):
        list1.append(re.sub(r'2021-11-29T12:', '', x[i]))
        list2.append(re.findall(r'\d{2}:\d{2}', list1[i]))
        list1[i] = list2[i][0]
    return list1


def sens():
    list1 = []
    for i in range(2, book2.max_column):
        list1.append(y_list(book2, i))
    for i in list1:
        i[0] = float(0)
    return list1


def plot_sens(T: list, S: list) -> list:
    """

    :rtype: list
    """
    for i in range(len(S)):
        plt.plot(T, S[i], label=f'{name_sens(book2, 1)[i]}')
    plt.xlabel("time", fontsize=14)
    plt.ylabel("value " + r'$\lambda$' + ", Нм")
    plt.title("Изменение длины волны со временем")
    ax = plt.gca()
    ax.tick_params(which="major",  # применяем основным делениям
                   length=5,  # длина делений
                   width=0.5,
                   pad=5,  # Расстояние между черточкой и ее подписью
                   labelsize=8,  # Размер подписи
                   bottom=True,
                   labelrotation=45)  # Поворот подписей
    ax.xaxis.set_major_locator(MultipleLocator(15))
    ax.yaxis.set_major_locator(MultipleLocator(0.25))
    # box = ax.get_position()
    # ax.set_position(box.x0, box.y0, box.width*0.8, box.height)
    plt.grid(color='k',
             linewidth='0.5',
             linestyle='--')
    plt.legend(loc="center left", bbox_to_anchor=(1, 0.5))
    plt.show()


def __print__(x):
    for i in x:
        print()
        for j in i:
            print(j, end=" ")


def strain(sensors_value):
    max_value = []
    sensor: list
    for sensor in sensors_value:
        max_ = sensor[0]
        # assert isinstance(sensor, list)
        for i in sensor:
            if abs(i) > abs(max_):
                max_ = i
        max_value.append(max_)
    # print(max_value)

    return max_value


def strain(sensors_value):
    max_value = []
    sensor: list
    for sensor in sensors_value:
        max_ = sensor[0]
        # assert isinstance(sensor, list)
        for i in sensor:
            if abs(i) > abs(max_):
                max_ = i
        max_value.append(max_)

    # print(max_value)

    def first_value():
        book1 = exel_file["track_from_2021-11-29_12-20-46"]
        first_val = name_sens(book1, 2)
        # print(first_val)
        return first_val

    def stress():
        stress_num = []
        for i in range(len(max_value)):
            stress_num.append(max_value[i] / (0.809 * float(first_value()[i])) * 1000)
        return stress_num

    # print(stress())


time = np.array((reg(y_list(book2, 1))))
# sensors = sens()
# print(type(sensors[5][9]))
sensors = np.array(sens())  # показание датчиков
strain(sensors)
# __print__(sensors)
# plot_sens(time, sensors)
